# Databricks notebook source
#!/usr/bin python3

# -------------------------------------------------------------------------
# Copyright (c) 2021 NHS England and NHS Improvement. All rights reserved.
# Licensed under the MIT License. See license.txt in the project root for
# license information.
# -------------------------------------------------------------------------

"""
FILE:           dbrks_shared_care_record_raw.py
DESCRIPTION:
                Databricks notebook with code to append new raw data to historical
                data for the NHSX Analyticus unit metrics within the Shared Care Record topic.
                
USAGE:
                ...
CONTRIBUTORS:   Craig Shenton, Mattia Ficarelli, Chris Todd
CONTACT:        data@nhsx.nhs.uk
CREATED:        19 Oct 2021
VERSION:        0.0.1
"""

# COMMAND ----------

# Install libs
# -------------------------------------------------------------------------
%pip install geojson==2.5.* tabulate requests pandas pathlib azure-storage-file-datalake  beautifulsoup4 numpy urllib3 lxml dateparser regex openpyxl pyarrow==5.0.*

# COMMAND ----------

# Imports
# -------------------------------------------------------------------------
# Python:
import os
import io
import tempfile
from datetime import datetime
import json
import regex as re

# 3rd party:
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from azure.storage.filedatalake import DataLakeServiceClient


# Connect to Azure datalake
# -------------------------------------------------------------------------
# !env from databricks secrets
CONNECTION_STRING = dbutils.secrets.get(scope="datalakefs", key="CONNECTION_STRING")

# COMMAND ----------

# MAGIC %run /Repos/prod/au-azure-databricks/functions/dbrks_helper_functions

# COMMAND ----------

# Load JSON config from Azure datalake
# -------------------------------------------------------------------------
file_path_config = "/config/pipelines/nhsx-au-analytics/"
file_name_config = "config_shared_care_record_dbrks.json"
file_system_config = "nhsxdatalakesagen2fsprod"
config_JSON = datalake_download(CONNECTION_STRING, file_system_config, file_path_config, file_name_config)
config_JSON = json.loads(io.BytesIO(config_JSON).read())

# COMMAND ----------

# Read parameters from JSON config
# -------------------------------------------------------------------------
file_system = config_JSON['pipeline']['adl_file_system']
source_path = config_JSON['pipeline']['raw']['source_path']
sink_path = config_JSON['pipeline']['raw']['sink_path']
historical_source_path = config_JSON['pipeline']['raw']['sink_path']

# COMMAND ----------

latestFolder = datalake_latestFolder(CONNECTION_STRING, file_system, source_path)
latestFolder

# COMMAND ----------

# Get list of all files in latest folder
def datalake_listDirectory(CONNECTION_STRING, file_system, source_path):
    try:
        service_client = DataLakeServiceClient.from_connection_string(CONNECTION_STRING)
        file_system_client = service_client.get_file_system_client(file_system=file_system)
        paths = file_system_client.get_paths(path=source_path)
        directory = []
        for path in paths:
            path.name = path.name.replace(source_path, '')
            directory.append(path.name)
    except Exception as e:
        print(e)
    return directory, paths
  
directory, paths = datalake_listDirectory(CONNECTION_STRING, file_system, source_path+latestFolder)
directory

# COMMAND ----------

from openpyxl import load_workbook


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


stp_df = pd.DataFrame()
trust_df = pd.DataFrame()
pcn_df = pd.DataFrame()
other_df = pd.DataFrame()

#loop through each submitted file in the landing area
for filename in directory:
    file = datalake_download(CONNECTION_STRING, file_system, source_path + latestFolder, filename)
    
    #Read current file inot an iobytes object, read that object and get list of sheet names
    sheets = get_sheetnames_xlsx(io.BytesIO(file))

    ### STP CALCULATIONS ###
    #list comprehension to get sheets with STP in the name from list of all sheets - ideally 1?
    sheet_name = [sheet for sheet in sheets if sheet.startswith("STP")]
    
    #Read current sheet. The output is a dictionary of columns
    xls_file = pd.read_excel(io.BytesIO(file), sheet_name=sheet_name, engine="openpyxl")
    for key in xls_file:
        #drop unnamed columns
        xls_file[key].drop(list(xls_file[key].filter(regex="Unnamed:")), axis=1, inplace=True)
        #remove empty rows
        xls_file[key] = xls_file[key].loc[~xls_file[key]["For Month"].isnull()]
        
        #rename columns based on order
        xls_file[key].rename(
            columns={
                list(xls_file[key])[0]: "For Month",
                list(xls_file[key])[1]: "ODS STP Code",
                list(xls_file[key])[2]: "STP Name",
                list(xls_file[key])[3]: "ICS Name (if applicable)",
                list(xls_file[key])[4]: "ShCR Programme Name",
                list(xls_file[key])[5]: "Name of ShCR System",
                list(xls_file[key])[6]: "Number of users with access to the ShCR",
                list(xls_file[key])[7]: "Number of citizen records available to users via the ShCR",
                list(xls_file[key])[8]: "Number of ShCR views in the past month",
                list(xls_file[key])[9]: "Number of unique user ShCR views in the past month",
                list(xls_file[key])[10]: "Completed by (email)",
                list(xls_file[key])[11]: "Date completed",
            },
            inplace=True,
        )
        
        # get excel file metadata
        STP_code = xls_file[key]["ODS STP Code"].unique()[0]  # get stp code for all sheets
        STP_name = xls_file[key]["STP Name"].unique()[0]  # get stp name for all sheets
        ICS_name = xls_file[key]["ICS Name (if applicable)"].unique()[0]  # get ics name for all sheets
          
        #For numeric fields, fill in blanks with zeros. Replace any non numeric entries with zero.
        xls_file[key]["Number of users with access to the ShCR"] = pd.to_numeric(xls_file[key]["Number of users with access to the ShCR"], errors='coerce').fillna(0).astype(int)
        xls_file[key]["Number of citizen records available to users via the ShCR"] = pd.to_numeric(xls_file[key]["Number of citizen records available to users via the ShCR"], errors='coerce').fillna(0).astype(int)
        xls_file[key]["Number of ShCR views in the past month"] = pd.to_numeric(xls_file[key]["Number of ShCR views in the past month"], errors='coerce').fillna(0).astype(int)
        xls_file[key]["Number of unique user ShCR views in the past month"] = pd.to_numeric(xls_file[key]["Number of unique user ShCR views in the past month"], errors='coerce').fillna(0).astype(int)

        # append results to dataframe dataframe
        stp_df = stp_df.append(xls_file[key], ignore_index=True)

    ### TRUST CALCULATIONS ###
    sheet_name = [sheet for sheet in sheets if sheet.startswith("Trust")]
    xls_file = pd.read_excel(io.BytesIO(file), sheet_name=sheet_name, engine="openpyxl")
    for key in xls_file:
        xls_file[key].drop(list(xls_file[key].filter(regex="Unnamed:")), axis=1, inplace=True)

        xls_file[key] = xls_file[key].loc[~xls_file[key]["For Month"].isnull()]
        xls_file[key].rename(
            columns={
                list(xls_file[key])[0]: "For Month",
                list(xls_file[key])[1]: "ODS Trust Code",
                list(xls_file[key])[2]: "Trust Name",
                list(xls_file[key])[3]: "Partner Organisation connected to ShCR?",
                list(xls_file[key])[4]: "Partner Organisation plans to be connected by Sept 2021?",
                list(xls_file[key])[5]: "Partner Organisation primary clinical system connect directly to the ShCR?",
            },
            inplace=True,
        )
        xls_file[key].insert(1, "ODS STP Code", STP_code, False)
        xls_file[key].insert(2, "STP Name", STP_name, False)
        xls_file[key].insert(3, "ICS Name (if applicable)", ICS_name, False)
        xls_file[key]["Partner Organisation connected to ShCR?"] = xls_file[key]["Partner Organisation connected to ShCR?"].map({"yes": 1, "no": 0, "Yes": 1, "No": 0}).fillna(0)   
        xls_file[key]["Partner Organisation primary clinical system connect directly to the ShCR?"] = xls_file[key]["Partner Organisation primary clinical system connect directly to the ShCR?"].map({"yes": 1, "no": 0, "Yes": 1, "No": 0}).fillna(0)   
        xls_file[key]["Partner Organisation plans to be connected by Sept 2021?"] = xls_file[key]["Partner Organisation plans to be connected by Sept 2021?"].map({"yes": 1, "no": 0, "Yes": 1, "No": 0}).fillna(0)        
        trust_df = trust_df.append(xls_file[key].iloc[:, 0:9], ignore_index=True)


    ### PCN CALCULATIONS ###
    sheet_name = [sheet for sheet in sheets if sheet.startswith("PCN")]
    xls_file = pd.read_excel(io.BytesIO(file), sheet_name=sheet_name, engine="openpyxl")
    
    for key in xls_file:
        #this isn't working properly - maybe switch to list of columns we DO want? 
        xls_file[key].drop(list(xls_file[key].filter(regex="Unnamed:")), axis=1, inplace=True)
        
        xls_file[key] = xls_file[key].loc[~xls_file[key]["For Month"].isnull()]
        xls_file[key].rename(
            columns={
                list(xls_file[key])[0]: "For Month",
                list(xls_file[key])[1]: "ODS PCN Code",
                list(xls_file[key])[2]: "PCN Name",
                list(xls_file[key])[3]: "Partner Organisation connected to ShCR?",
                list(xls_file[key])[4]: "Partner Organisation plans to be connected by Sept 2021?",
                list(xls_file[key])[5]: "Partner Organisation primary clinical system connect directly to the ShCR?",
            },
            inplace=True,
        )
        xls_file[key].insert(1, "ODS STP Code", STP_code, False)
        xls_file[key].insert(2, "STP Name", STP_name, False)
        xls_file[key].insert(3, "ICS Name (if applicable)", ICS_name, False)
        xls_file[key]["Partner Organisation connected to ShCR?"] = xls_file[key]["Partner Organisation connected to ShCR?"].map({"yes": 1, "no": 0, "Yes": 1, "No": 0}).fillna(0)   
        xls_file[key]["Partner Organisation primary clinical system connect directly to the ShCR?"] = xls_file[key]["Partner Organisation primary clinical system connect directly to the ShCR?"].map({"yes": 1, "no": 0, "Yes": 1, "No": 0}).fillna(0)   
        xls_file[key]["Partner Organisation plans to be connected by Sept 2021?"] = xls_file[key]["Partner Organisation plans to be connected by Sept 2021?"].map({"yes": 1, "no": 0, "Yes": 1, "No": 0}).fillna(0)   

        pcn_df = pcn_df.append(xls_file[key], ignore_index=True)
        
#     Other calculations
#     sheet_name = [sheet for sheet in sheets if sheet.startswith('Other')]
#     xls_file = pd.read_excel(io.BytesIO(file), sheet_name=sheet_name, engine='openpyxl')
#     for key in xls_file:
#     xls_file[key].drop(list(xls_file[key].filter(regex = 'Unnamed:')), axis = 1, inplace = True)
#     xls_file[key] = xls_file[key].loc[~xls_file[key]["For Month"].isnull()]
#     xls_file[key].rename(columns={list(xls_file[key])[0]:"For Month",
#                                  list(xls_file[key])[1]:"ODS STP Code",
#                                  list(xls_file[key])[2]:"ODS PCN Code",
#                                  list(xls_file[key])[2]:"Other partner",
#                                  list(xls_file[key])[3]:"Partner Organisation connected to ShCR?",
#                                  list(xls_file[key])[4]:"Partner Organisation plans to be connected by Sept 2021?",
#                                  list(xls_file[key])[5]:"Partner Organisation's primary clinical system connect directly to the ShCR?"},inplace=True)
#     xls_file[key].drop('ODS STP Code', axis=1, inplace=True)
#     ls_file[key].insert(1, "ODS STP Code", STP_code, False)
#     xls_file[key].insert(2, "STP Name", STP_name, False)
#     xls_file[key].insert(3, "ICS Name (if applicable)", ICS_name, False)
#     print(list(xls_file[key]))
#     other_df = other_df.append(xls_file[key], ignore_index=True)
    
#Remove any non-required columns from final output
pcn_df = pcn_df[['For Month', 'ODS STP Code', 'STP Name', 'ICS Name (if applicable)', 'ODS PCN Code', 'PCN Name', 'Partner Organisation connected to ShCR?', 'Partner Organisation plans to be connected by Sept 2021?', 'Partner Organisation primary clinical system connect directly to the ShCR?']]

trust_df = trust_df[['For Month', 'ODS STP Code', 'STP Name', 'ICS Name (if applicable)', 'ODS Trust Code', 'Trust Name', 'Partner Organisation connected to ShCR?', 'Partner Organisation plans to be connected by Sept 2021?', 'Partner Organisation primary clinical system connect directly to the ShCR?']]

stp_df= stp_df[['For Month', 'ODS STP Code', 'STP Name', 'ICS Name (if applicable)', 'ShCR Programme Name', 'Name of ShCR System', 'Number of users with access to the ShCR', 'Number of citizen records available to users via the ShCR', 'Number of ShCR views in the past month', 'Number of unique user ShCR views in the past month', 'Completed by (email)', 'Date completed']]

# COMMAND ----------

folder_date = pd.to_datetime(latestFolder) - pd.DateOffset(months=1)

stp_df['For Month'] = folder_date
stp_df['Date completed'] = pd.to_datetime(stp_df['Date completed'],errors='coerce')

pcn_df['For Month'] = folder_date
trust_df['For Month'] = folder_date

# COMMAND ----------

### check for duplicates
# import collections
# a = [item for item, count in collections.Counter(trust_df['ODS Trust Code']).items() if count > 1]
# trust_df[trust_df['ODS Trust Code'].isin(a)].sort_values(by='ODS Trust Code')

# #a = [item for item, count in collections.Counter(pcn_df['ODS PCN Code']).items() if count > 1]
# #pcn_df[pcn_df['ODS PCN Code'].isin(a)].sort_values(by='ODS Trust Code')

# COMMAND ----------

# #cast all date fields to datetime format and set to 1st of the month
# dt =lambda dt: dt.replace(day=1)

# pcn_df['For Month'] = pd.to_datetime(pcn_df['For Month']).apply(dt)
# stp_df['For Month'] = pd.to_datetime(stp_df['For Month']).apply(dt)
# trust_df['For Month'] = pd.to_datetime(trust_df['For Month']).apply(dt)

# COMMAND ----------

##Calculate aggregate numbers for Trusts
trust_count_df = trust_df.groupby('STP Name')['Partner Organisation connected to ShCR?'].size().reset_index(name='Total')
trust_count_df_2 = trust_df.groupby('STP Name')['Partner Organisation connected to ShCR?'].sum().reset_index(name='Total')
trust_count_df['Number Connected'] = trust_count_df_2['Total']
trust_count_df['Percent'] = trust_count_df_2['Total']/trust_count_df['Total']
trust_count_df['Type'] = 'Trust'

# COMMAND ----------

##Calculate aggregate numbers for PCNs
pcn_count_df = pcn_df.groupby('STP Name')['Partner Organisation connected to ShCR?'].size().reset_index(name='Total')
pcn_count_df2 = pcn_df.groupby('STP Name')['Partner Organisation connected to ShCR?'].sum().reset_index(name='Total')
pcn_count_df['Number Connected'] = pcn_count_df2['Total']
pcn_count_df['Percent'] = pcn_count_df2['Total']/pcn_count_df['Total']
pcn_count_df['Type'] = 'PCN'

# COMMAND ----------

#SNAPSHOT SUMMARY
#Write pages to Excel file in iobytes
files = [stp_df, trust_df, trust_count_df, pcn_df, pcn_count_df]
sheets = ['STP', 'Trust', 'Trust Count', 'PCN', 'PCN Count']
excel_sheet = io.BytesIO()

writer = pd.ExcelWriter(excel_sheet, engine='openpyxl')
for count, file in enumerate(files):
    file.to_excel(writer, sheet_name=sheets[count], index=False)
writer.save()

# COMMAND ----------

#SNAPSHOT SUMMARY
#Send Excel snapshot File to test Output in datalake
current_date_path = datetime.now().strftime('%Y-%m-%d') + '/'

file_contents = excel_sheet
datalake_upload(file_contents, CONNECTION_STRING, "nhsxdatalakesagen2fsprod", "proc/projects/nhsx_slt_analytics/shcr/excel_summary/"+current_date_path, "excel_output.xlsx")

# COMMAND ----------

#Pull historical files
latestFolder = datalake_latestFolder(CONNECTION_STRING, file_system, historical_source_path)

file_name_list = datalake_listContents(CONNECTION_STRING, file_system, historical_source_path+latestFolder)
for file in file_name_list:
  if 'stp' in file:
    stp_df_historic = datalake_download(CONNECTION_STRING, file_system, historical_source_path+latestFolder, file)
    stp_df_historic = pd.read_parquet(io.BytesIO(stp_df_historic), engine="pyarrow")
    stp_df_historic['For Month'] = pd.to_datetime(stp_df_historic['For Month'])
    stp_df_historic['Date completed'] = pd.to_datetime(stp_df_historic['Date completed'],errors='coerce')
  if 'pcn' in file:
    pcn_df_historic = datalake_download(CONNECTION_STRING, file_system, historical_source_path+latestFolder, file)
    pcn_df_historic = pd.read_parquet(io.BytesIO(pcn_df_historic), engine="pyarrow")
    pcn_df_historic['For Month'] = pd.to_datetime(pcn_df_historic['For Month'])
  if 'trust' in file:
    trust_df_historic = datalake_download(CONNECTION_STRING, file_system, historical_source_path+latestFolder, file)
    trust_df_historic = pd.read_parquet(io.BytesIO(trust_df_historic), engine="pyarrow")
    trust_df_historic['For Month'] = pd.to_datetime(trust_df_historic['For Month'])
    

# COMMAND ----------

# # Append new data to historical data
# # -----------------------------------------------------------------------
#STP
dates_in_historic = stp_df_historic["For Month"].unique().tolist()
dates_in_new = stp_df["For Month"].unique().tolist()[0]

if dates_in_new in dates_in_historic:
  print('STP Data already exists in historical STP data')
else:
  stp_df_historic = stp_df_historic.append(stp_df)
  stp_df_historic = stp_df_historic.sort_values(by=['For Month'])
  stp_df_historic = stp_df_historic.reset_index(drop=True)
  stp_df_historic = stp_df_historic.astype(str)
  
#PCN
dates_in_historic = pcn_df_historic["For Month"].unique().tolist()
dates_in_new = pcn_df["For Month"].unique().tolist()[0]

if dates_in_new in dates_in_historic:
  print('PCN Data already exists in historical STP data')
else:
  pcn_df_historic = pcn_df_historic.append(pcn_df)
  pcn_df_historic = pcn_df_historic.sort_values(by=['For Month'])
  pcn_df_historic = pcn_df_historic.reset_index(drop=True)
  pcn_df_historic = pcn_df_historic.astype(str)
  
#TRUST
dates_in_historic = trust_df_historic["For Month"].unique().tolist()
dates_in_new = trust_df["For Month"].unique().tolist()[0]

if dates_in_new in dates_in_historic:
  print('Trust Data already exists in historical STP data')
else:
  trust_df_historic = trust_df_historic.append(trust_df)
  trust_df_historic = trust_df_historic.sort_values(by=['For Month'])
  trust_df_historic = trust_df_historic.reset_index(drop=True)
  trust_df_historic = trust_df_historic.astype(str)

# COMMAND ----------

# Upload processed data to datalake
current_date_path = datetime.now().strftime('%Y-%m-%d') + '/'
file_contents = io.BytesIO()
#pcn
pcn_df_historic.to_parquet(file_contents, engine="pyarrow")
datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_pcn_data_month_count.parquet")
# #stp
file_contents = io.BytesIO()
stp_df_historic.to_parquet(file_contents, engine="pyarrow")
datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_stp_data_month_count.parquet")
# #trust
file_contents = io.BytesIO()
trust_df_historic.to_parquet(file_contents, engine="pyarrow")
datalake_upload(file_contents, CONNECTION_STRING, file_system, sink_path+current_date_path, "shcr_partners_trust_data_month_count.parquet")
